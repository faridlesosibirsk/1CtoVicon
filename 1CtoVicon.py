import openpyxl
import datetime
import win32api
import readchar

try:
    excel_file = openpyxl.load_workbook('Прием на обучение на бакалавриат-специалитет 2021 (ЛФ).xlsx')
    employees_sheet = excel_file['Прием на обучение на бакалавриа']
except:
    win32api.MessageBox(0, 'Прием на обучение на бакалавриат-специалитет 2021 (ЛФ).xlsx Лист:Прием на обучение на бакалавриа', 'Ожидается файл')
    #alert("Прием на обучение на бакалавриат-специалитет 2021 (ЛФ).xlsx/nПрием на обучение на бакалавриа");


class Abiturient:
    number = 0
    fio = ""
    preemptionRight = ""
    totalPoints = ""
    totalPointsSubjects = ""
    totalPointsCompetitive = ""
    typeEducationDocument = ""
    order = ""
    submissionOrder = ""
    consentTransfer = ""
    needHostel = ""
    applicationNumber = ""
    dateApplication = ""
    formTraining = ""
    sourceFunding = ""
    withoutEntranceTests = ""
    targetValue = ""
    typeEducation = ""
    original = ""
    credited = ""
    expelled = ""
    creditedOtherDirection = ""
    enrolledOtherCompetitivePlacesConditions = ""
    uniqueCode = ""
    specialRight = ""

class Welcome:
    dateFormation = datetime.datetime.now()
    timeFormation = datetime.datetime.now()
    competitionGroup = ""
    formTraining = ""
    trainingLevel = ""
    fieldStudySpecialty = ""
    basisReceipt = ""
    sourceFunding = ""
    receptionCategory = ""
    totalSeats = ""
    totalCredited = ""
    toBeCredited = ""
    #abiturient = Abiturient()
    abiturient = []
    #count = 0

welcome = Welcome()
for x in range(1, employees_sheet.max_row+1):
    line = employees_sheet.cell(row=x, column=1).value
    if line != None and isinstance(line, str):
        if "Дата" in line:
            welcome.dateFormation = line.split('Дата формирования - ')[1].split('. ')[0]
            #print(welcome.dateFormation)
            #print(employees_sheet.cell(row=x, column=1).value)
        if "Время" in line:
            welcome.timeFormation = line.split('Время формирования - ')[1].split('.')[0]
            #print(welcome.timeFormation)
        if "Конкурсная" in line:
            welcome.competitionGroup = line.split('Конкурсная группа - ')[1]
            print(welcome.competitionGroup)
        if "Форма" in line:
            welcome.formTraining = line.split('Форма обучения - ')[1]
            #print(welcome.formTraining)
        if "Уровень" in line:
            welcome.trainingLevel = line.split('Уровень подготовки - ')[1]
            #print(welcome.trainingLevel)
        if "УГС" in line:
            welcome.fieldStudySpecialty = line.split('УГС/Направление подготовки/специальность - ')[1]
            #print(welcome.fieldStudySpecialty)
        if "Основание" in line:
            welcome.basisReceipt = line.split('Основание поступления - ')[1]
            #print(welcome.basisReceipt)
        if "Источник" in line:
            welcome.sourceFunding = line.split('Источник финансирования - ')[1]
            #print(welcome.sourceFunding)
        if "Категория" in line:
            welcome.receptionCategory = line.split('Категория приема - ')[1]
            #print(welcome.receptionCategory)
        if "Всего" in line:
            welcome.totalSeats = line.split('Всего мест: ')[1].split('. ')[0]
            #print(welcome.totalSeats)
        if "Зачислено" in line:
            welcome.totalCredited = line.split('Зачислено: ')[1].split('. ')[0]
            #print(welcome.totalCredited)
        if "зачислению" in line:
            welcome.toBeCredited = line.split('К зачислению: ')[1].split('.')[0]
            #print(welcome.toBeCredited)            
    elif line != None and isinstance(line, int): 
        abit = []
        for y in range(1, employees_sheet.max_column+1):
            field = employees_sheet.cell(row=x, column=y).value
            abit.append(field)
        welcome.abiturient.append(abit)
        #print(len(welcome.abiturient))

    elif  (( line == None ) and (len(welcome.abiturient) > 0)) or (x == employees_sheet.max_row):
    #elif len(welcome.abiturient) > 0 or x == employees_sheet.max_row:
        count = 0
        #for i in range( len(welcome.abiturient)):
        #    for y in range(employees_sheet.max_column):
        #        print(welcome.abiturient[i][y], end=" ")
        #    print("\n")
        print( len(welcome.abiturient) )
        print( '\n' )
        count = len(welcome.abiturient)    
        for i in range( count ):   
            one = [None]*46
            temp = [None]*25
            for y in range(employees_sheet.max_column):
                temp[y] = welcome.abiturient[i-1][y]
            one[0] = temp[1] #ФИО
            one[2] = temp[23] #Уникальный код
            #one[3] = temp[17] #Абитуриент имеет аттестат/диплом
            if "Среднее специальное" in temp[17]: one[3] = 3
            if "Начальное профессиональное" in temp[17]: one[3] = 3
            if "Среднее общее образование" in temp[17]: one[3] = 2
            one[11] = temp[11] #Номер заявления
            #one[14] = temp[6] #Документ об образовании
            if "Копия" in temp[6]: one[14] = 0
            if "Оригинал" in temp[6]: one[14] = 1
            #one[15] = temp[9] #Согласие на зачисление
            if temp[9] == None: one[15] = 0
            if temp[9] != None: one[15] = 1
            #one[16] = temp[10] #Нуждается в общежитии
            if temp[10] == None: one[16] = 0
            if temp[10] != None: one[16] = 1
            #one[17] = temp[12] #Дата регистрации заявления
            one[17] = temp[12]
            #one[18] = temp[18] # [18-21] Статус зачисления
            one[18] = 1
            if temp[19] != None: one[18] = 2
            if temp[20] != None: one[18] = 3
            if temp[21] != None: one[18] = 4
            if temp[22] != None: one[18] = 4
            #one[19] = temp[13] #Форма обучения
            if "Заочная" in temp[13]: one[19] = 3
            if "Очная" in temp[13]: one[19] = 1
            if "Очно-заочная" in temp[13]: one[19] = 2
            #one[20] = temp[14] #Источник финансирования
            if "Федеральный бюджет" in temp[14]: one[20] = 1
            if "Внебюджетные средства" in temp[14]: one[20] = 5
            #one[21] = temp[14] #Уровень бюджета
            one[21] = 1
            #if "Федеральный бюджет" in temp[14]: one[21] = 1
            #one[26] = temp[16] #поступает без вступительных испытаний
            one[26] = 1 
            #if temp[16] != None: one[26] = 0
            #if temp[16] == None: one[26] = 1

            try:
                if i == 0:
                    template_file = openpyxl.load_workbook('template_abiturs.xlsx')
                else:
                    template_file = openpyxl.load_workbook(welcome.competitionGroup + '.xlsx')
                template_sheet = template_file['перечень абитуриентов']
            except:
                win32api.MessageBox(0, 'template_abiturs.xlsx/nперечень абитуриентов', 'Ожидается файл')
                #alert("template_abiturs.xlsx Лист:перечень абитуриентов");

            template_sheet.append(one)
            #print(one)
            template_file.save(welcome.competitionGroup + '.xlsx')
            #del one[:]
        for i in range( count):
            willDel = welcome.abiturient.pop()
        #print( len(welcome.abiturient) )
        #welcome.abiturient = []*1
        #print(welcome.competitionGroup)
#del welcome

print("Нажмите любую клавишу для выхода.")
k = readchar.readchar()

# sheet names
#print(excel_file.sheetnames)

#employees_sheet = excel_file['1c']
#print(f'Total Rows = {employees_sheet.max_row} and Total Columns = {employees_sheet.max_column}')

#for x in range(1, employees_sheet.max_row+1):
#    print(employees_sheet.cell(row=x, column=1).value)

#for x in range(1, employees_sheet.max_column+1):
#    print(employees_sheet.cell(row=1, column=x).value)
